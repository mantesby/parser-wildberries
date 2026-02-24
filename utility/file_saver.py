import pandas as pd
import logging
import openpyxl

class XlsxManager():
    def __init__(self) -> None:
        self.logger = logging.getLogger("filename")
        
    def _images_text(self, pics):
        text = ""
        for i in pics:
            text += f"{i},"
        return text
    
    def _sizes_text(self, sizes):
        text = ""
        for i in sizes:
            text += f"{i},"
        return text
        
    def _main_info_text(self, info):
        text = ""
        for i in info:
            text += f"{i["name"]}: {i["value"]}"
        return text
    
    def _additional_info_text(self, info):
        text = ""
        for i in info:
            text += f"{i["name"]}: {i["value"]}"
        return text
    
    def _add_to_data(self, data_for_excel: dict, row_init):
        images_text = self._images_text(row_init["pics"])
        sizes_text = self._sizes_text(row_init["sizes"])
        main_text = self._main_info_text(row_init["main_info"]["options"])
        additional_text = self._additional_info_text(row_init["additional_info"]["options"])
                
        data_for_excel["Артикул"].append(row_init["articule"])
        data_for_excel["Ссылка на товар"].append(row_init["url_card"])
        data_for_excel["Название"].append(row_init["name"])
        data_for_excel["Цена"].append(row_init["price"])
        data_for_excel["Описание"].append(row_init["description"])
        data_for_excel["Ссылки на изображения"].append(images_text)
        data_for_excel["Размеры товара"].append(sizes_text)
        data_for_excel["Основная информация"].append(main_text)
        data_for_excel["Дополнительная информация"].append(additional_text)
        data_for_excel["Название селлера"].append(row_init["seller"])
        data_for_excel["Ссылка на селлера"].append(row_init["page_seller"])
        data_for_excel["Количество отзывов"].append(row_init["reviews"])
        data_for_excel["Остатки товара"].append(row_init["total"])
        data_for_excel["Рейтинг"].append(row_init["reviewRating"])
        
    def _get_country(self, row_init):
        for i in row_init["additional_info"]["options"]:
                if i["name"] == "Страна производства":
                    return i["value"]
        
    def convert(self, row_data_list):
        self.logger.debug("Подготовка данных в формат для excel")
        data_for_excel = {
            "Ссылка на товар": [],
            "Артикул": [],
            "Название": [],
            "Цена": [],
            "Описание": [],
            "Ссылки на изображения": [],
            "Основная информация": [],
            "Дополнительная информация": [],
            "Название селлера": [],
            "Ссылка на селлера": [],
            "Размеры товара": [],
            "Остатки товара": [],
            "Рейтинг": [],
            "Количество отзывов": []
        }
        filter_data_for_excel = {
            key: [] for key in data_for_excel
        }
        
        for row_init in row_data_list:
            self._add_to_data(data_for_excel, row_init)
            
            country = self._get_country(row_init)
            if row_init["price"] <= 10000 and row_init["reviewRating"] >= 4.5 and country == "Россия":
                
                self._add_to_data(filter_data_for_excel, row_init)
        self.logger.debug("Данные успешно подготовлены")
        return data_for_excel, filter_data_for_excel
            
        
    def save_to_excel(self, data):
        filename = "parser_wildberries.xlsx"
        filename_filter = "parses_wildberries_filter.xlsx"
        data_excel, filter_data_excel = self.convert(data)
        df = pd.DataFrame(data_excel)
        df.to_excel(filename, index=False)
        book = openpyxl.load_workbook(filename)
        sheet = book["Sheet1"]
        start_row = sheet.max_row
        
        df2 = pd.DataFrame(filter_data_excel)
        df2.to_excel(filename_filter, index=False)
        book2 = openpyxl.load_workbook(filename_filter)
        sheet2 = book2["Sheet1"]
        start_row2 = sheet2.max_row
        with pd.ExcelWriter(filename, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            df.to_excel(
                writer,
                sheet_name="Sheet1",
                index=False,
                header=False,      
                startrow=start_row
            )
        self.logger.info("Добавлены данные в обычный файл xlsx")
            
        with pd.ExcelWriter(filename_filter, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            df2.to_excel(
                writer,
                sheet_name="Sheet1",
                index=False,
                header=False,      
                startrow=start_row2  
            )
        self.logger.info("Добавлены данные в фильтрованный файл xlsx")
        